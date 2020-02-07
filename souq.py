import csv
import re

import lxml
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

excel_filename = 'data.xlsx'

webpages = [{
    'link': 'https://deals.souq.com/eg-en/mobiles/cc/482',
    'sheet_name': 'Mobiles',
    'header': ['Mobile', 'Price (EGP)'],
    'active': True
}, {
    'link': 'https://deals.souq.com/eg-en/health-and-beauty/cc/428',
    'sheet_name': 'Health & Beauty',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/mobiles-and-electronics/cc/424',
    'sheet_name': 'Electronics',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/moms-and-babies/cc/351',
    'sheet_name': 'Moms & Babies',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/home-and-kitchen/cc/493',
    'sheet_name': 'Home & Kitchen',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/stationery/cc/402',
    'sheet_name': 'Office Products',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/car-accessories/cc/400',
    'sheet_name': 'Automotive',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/appliances/cc/348',
    'sheet_name': 'Appliances',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/sports/cc/403',
    'sheet_name': 'Sports',
    'header': ['Name', 'Price (EGP)'],
}, {
    'link': 'https://deals.souq.com/eg-en/toys/cc/495',
    'sheet_name': 'Toys',
    'header': ['Name', 'Price (EGP)'],
}]


def crawl(link):
    page = requests.get(link)
    page_soup = BeautifulSoup(page.text, 'lxml')

    # Get items titles
    page_items = [
        i.getText().strip() for i in page_soup.select("span.itemTitle a")
    ]

    # Get items prices
    # Use regex to extract numbers only
    page_prices = [
        re.search("(^[\d,.]*)", i.getText()).group(0)
        for i in page_soup.select("h5.price span.is")
    ]

    # Zip items and prices in a list and return it
    page_pairs = list(zip(page_items, page_prices))
    return page_pairs


def write_sheet(wb, sheet_name, header, pairs, active=False):
    """ Write an Excel Sheet """
    ws = None

    if active:
        # If it's the first sheet to write
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    # Write headers
    ws.cell(1, 1, header[0])
    ws.cell(1, 2, header[1])

    # Write data
    for row in range(2, len(pairs) + 1):
        for col in range(1, 3):
            _ = ws.cell(column=col, row=row, value=f"{pairs[row-2][col-1]}")


# Create new Excel file
wb = Workbook()

# Loop through webpages, get each one's data, and write it to the excel sheet
for page in webpages:
    print("Getting {}... ".format(page['sheet_name']), end='', flush=True)
    content = crawl(page['link'])
    print("Done.")

    write_sheet(
        wb,
        page['sheet_name'],
        page['header'],
        content,
        active=page.get('active') or False)

# Save the final excel file
wb.save(filename=excel_filename)

print("Data saved to '{}'".format(excel_filename))